from pathlib import Path
from typing import List
import pandas as pd
from openpyxl import Workbook
from services.reference_book import ReferenceBook


class FileProcessor:
    """Сервис для обработки Excel файлов с кодами"""

    @staticmethod
    def extract_article(filename: str) -> str:
        """
        Извлекает артикул из названия файла.
        Артикул — первое слово до первого пробела.
        """
        return filename.split()[0].strip()

    @staticmethod
    def read_codes_from_file(file_path: Path) -> List[str]:
        df = pd.read_excel(file_path, header=None, skiprows=1, dtype=str, engine="openpyxl")

        if df.shape[1] < 2:
            return []

        codes = df.iloc[:, 1].str.strip().dropna()
        codes = codes[codes != ""].tolist()

        return codes

    @staticmethod
    def create_result_file(
        barcode: str,
        codes: List[str],
        output_path: Path
    ) -> None:
        """
        Создает результирующий Excel файл с указанной структурой.
        
        Структура файла:
        - Строка 1, столбец A: "коды"
        - Строка 2, столбец A: штрихкод
        - Строки 3+, столбец A: коды из входного файла
        
        Args:
            article: Артикул (используется для названия файла)
            barcode: Штрихкод для записи во вторую строку
            codes: Список кодов для записи с третьей строки
            output_path: Путь для сохранения файла
        """
        wb = Workbook()
        ws = wb.active
        
        # Первая строка: "коды"
        ws['A1'] = "коды"
        
        # Вторая строка: штрихкод
        ws['A2'] = barcode
        
        # Третья и последующие строки: коды
        for idx, code in enumerate(codes, start=3):
            ws[f'A{idx}'] = code
        
        # Сохраняем файл
        wb.save(output_path)

    @classmethod
    def process_file(
        cls,
        input_file_path: Path,
        output_dir: Path,
        filename: str
    ) -> tuple[Path, str]:
        """
        Обрабатывает входящий файл и создает результирующий файл.
        
        Args:
            input_file_path: Путь к входящему файлу
            output_dir: Директория для сохранения результата
            filename: Оригинальное название файла
            
        Returns:
            tuple: (Путь к результирующему файлу, артикул)
            
        Raises:
            ValueError: При различных ошибках валидации
        """
        # 1. Извлекаем артикул из названия файла
        article = cls.extract_article(filename)
        if not article:
            raise ValueError("Не удалось извлечь артикул из названия файла")
        
        # 2. Ищем штрихкод в справочнике
        barcode = ReferenceBook.get_barcode(article)
        if not barcode:
            raise ValueError(f"Артикул «{article}» (Файл {filename}) не найден в справочнике")
        
        # 3. Читаем коды из файла
        codes = cls.read_codes_from_file(input_file_path)
        if not codes:
            raise ValueError(f"В файле «{filename}» не найдено кодов в столбце B")
        
        # 4. Создаем результирующий файл
        output_filename = f"codes_{article}.xlsx"
        output_path = output_dir / output_filename
        
        cls.create_result_file(barcode, codes, output_path)
        
        return output_path, article
